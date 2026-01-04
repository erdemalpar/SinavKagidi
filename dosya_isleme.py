"""
Dosya İşleme Modülü - Akıllı Format Algılama
Herhangi bir formattaki soru dosyasını otomatik algılar ve işler
"""

import os
import re
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
from PIL import Image
import json


def akilli_soru_ayikla(metin):
    """
    Herhangi bir formattaki soruyu akıllıca algılar ve çıkarır
    Format bağımsız, otomatik tespit yapılır
    """
    try:
        metin = metin.strip()
        if not metin or len(metin) < 10:
            return None
        
        # Satırlara ayır ve boş olanları temizle
        satirlar = [s.strip() for s in metin.split('\n') if s.strip()]
        
        if len(satirlar) < 2:
            return None
        
        # Cevap satırını bul ve kaldır
        dogru_cevap = 'A'
        cevap_indeksi = -1
        
        for i, satir in enumerate(satirlar):
            # Cevap pattern'leri: "Cevap: A", "Doğru: A", "Answer: A", "Cevap A", vs.
            cevap_match = re.search(r'(?:Cevap|Doğru|Yanıt|Answer|Cvp)[:\s]*([A-Ea-e])', satir, re.IGNORECASE)
            if cevap_match:
                dogru_cevap = cevap_match.group(1).upper()
                cevap_indeksi = i
                # Eğer cevap satırında şık metni de varsa onu al
                cevap_metni = re.sub(r'(?:Cevap|Doğru|Yanıt|Answer|Cvp)[:\s]*[A-Ea-e]\)?\s*', '', satir, flags=re.IGNORECASE).strip()
                break
        
        # Cevap satırını listeden çıkar
        if cevap_indeksi >= 0:
            satirlar.pop(cevap_indeksi)
        
        # Şıkları akıllıca tespit et
        siklar = {}
        soru_satirlari = []
        sik_basladi = False
        
        harfler = ['A', 'B', 'C', 'D', 'E']
        sik_sayaci = 0
        
        for satir in satirlar:
            # Puan bilgisi içeren satırları atla
            if re.search(r'\(\s*\d+\s*puan\s*\)', satir, re.IGNORECASE):
                if not sik_basladi:
                    soru_satirlari.append(satir)
                continue
            
            # Şık pattern'leri çok çeşitli olabilir:
            # - "A) Metin", "A. Metin", "A- Metin", "A: Metin"
            # - "a) Metin", "a. Metin"
            # - "(A) Metin", "[A] Metin"
            sik_match = re.match(r'^[\(\[]?([A-Ea-e])[\)\]\.\-:\s]+(.+)$', satir)
            
            if sik_match:
                sik_basladi = True
                harf = sik_match.group(1).upper()
                icerik = sik_match.group(2).strip()
                siklar[harf] = icerik
                sik_sayaci += 1
            elif sik_basladi:
                # Şıklar başladıysa ve harf yoksa, önceki şıkkın devamı olabilir
                if siklar:
                    son_harf = list(siklar.keys())[-1]
                    siklar[son_harf] += ' ' + satir
            else:
                # Henüz şık başlamadıysa, soru metni
                soru_satirlari.append(satir)
        
        # Eğer hiç harfli şık bulunamadıysa, satırları sırayla şık olarak kabul et
        if not siklar and len(satirlar) > 2:
            soru_satirlari = [satirlar[0]]
            for i, satir in enumerate(satirlar[1:6]):  # Maksimum 5 şık
                if i < len(harfler):
                    # Numaralı olabilir: "1. Metin", "1) Metin"
                    numara_match = re.match(r'^\d+[\).:\s]+(.+)$', satir)
                    if numara_match:
                        siklar[harfler[i]] = numara_match.group(1).strip()
                    else:
                        siklar[harfler[i]] = satir
        
        # Soru metni
        soru_metni = ' '.join(soru_satirlari).strip()
        
        # Puan bilgisini temizle
        soru_metni = re.sub(r'\(\s*\d+\s*puan\s*\)', '', soru_metni, flags=re.IGNORECASE).strip()
        
        # En az 2 şık olmalı
        if len(siklar) < 2:
            print(f"⚠ Yetersiz şık: {len(siklar)} şık bulundu")
            return None
        
        # Cevap metni varsa ve cevap şıkkı eksikse ekle
        if 'cevap_metni' in locals() and cevap_metni and dogru_cevap not in siklar:
            siklar[dogru_cevap] = cevap_metni
        
        return {
            'soru_metni': soru_metni,
            'secenek_a': siklar.get('A', ''),
            'secenek_b': siklar.get('B', ''),
            'secenek_c': siklar.get('C', ''),
            'secenek_d': siklar.get('D', ''),
            'secenek_e': siklar.get('E', ''),
            'dogru_cevap': dogru_cevap,
            'konu': '',
            'zorluk': 'Orta',
            'puan': 5
        }
    
    except Exception as e:
        print(f"❌ Soru ayıklama hatası: {e}")
        import traceback
        traceback.print_exc()
        return None


def pdf_isle(dosya_yolu):
    """PDF dosyasından soru çıkarır"""
    sorular = []
    
    try:
        reader = PdfReader(dosya_yolu)
        tum_metin = ""
        
        for sayfa in reader.pages:
            tum_metin += sayfa.extract_text() + "\n"
        
        # Soruları farklı pattern'lerle ayır
        patterns = [
            r'\n\s*(?:S|Soru)\s*(\d+)[:.)\s]+',  # S1., Soru 1:
            r'\n\s*(\d+)[.)]\s+',  # 1., 1)
            r'\n\s*Question\s*(\d+)[:.)\s]+',  # Question 1:
        ]
        
        soru_bloklari = None
        for pattern in patterns:
            bloklari = re.split(pattern, tum_metin, flags=re.IGNORECASE)
            if len(bloklari) > 2:
                soru_bloklari = bloklari
                break
        
        if soru_bloklari:
            i = 1
            while i < len(soru_bloklari) - 1:
                soru_numarasi = soru_bloklari[i]
                soru_metni_blok = soru_bloklari[i + 1]
                
                soru = akilli_soru_ayikla(soru_metni_blok)
                if soru:
                    print(f"✓ PDF Soru {soru_numarasi} çıkarıldı")
                    sorular.append(soru)
                
                i += 2
    
    except Exception as e:
        print(f"❌ PDF işleme hatası: {e}")
        return []
    
    return sorular


def docx_isle(dosya_yolu):
    """DOCX dosyasından soru çıkarır - Otomatik format algılama"""
    sorular = []
    
    try:
        doc = Document(dosya_yolu)
        tum_metin = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        
        print(f"📄 DOCX toplam metin uzunluğu: {len(tum_metin)} karakter")
        
        # Farklı soru ayırma pattern'lerini dene
        patterns = [
            (r'\n\s*(?:S|Soru|Question|K\.)\s*(\d+)[\.\)\:\-\s]+', 'Soru 1: / S.1'),
            (r'\n\s*(\d+)[\.\)\:\-]\s+', '1. / 1) / 1- / 1:'), 
            (r'\n\s*(\d+)\.\s*Soru', '1. Soru'),
            # Eğer numaralandırma yoksa (Word otomatik liste), paragraf başlarını vs. algılamak zor.
            # Şimdilik numaralandırmaya güveniyoruz.
        ]
        
        soru_bloklari = None
        kullanilan_pattern = None
        
        for pattern, aciklama in patterns:
            bloklari = re.split(pattern, tum_metin, flags=re.IGNORECASE)
            if len(bloklari) > 2:
                soru_bloklari = bloklari
                kullanilan_pattern = aciklama
                print(f"✓ Soru formatı tespit edildi: {aciklama}")
                break
        
        if not soru_bloklari:
            print("⚠ Standart soru numaralandırması bulunamadı, tüm metin tek soru olarak işleniyor")
            soru = akilli_soru_ayikla(tum_metin)
            if soru:
                sorular.append(soru)
            return sorular
        
        print(f"📋 Bulunan soru blok sayısı: {(len(soru_bloklari) - 1) // 2}")
        
        # Blokları işle
        i = 1
        while i < len(soru_bloklari) - 1:
            soru_numarasi = soru_bloklari[i]
            soru_metni_blok = soru_bloklari[i + 1]
            
            print(f"\n--- Soru {soru_numarasi} işleniyor ---")
            print(f"Blok uzunluğu: {len(soru_metni_blok)} karakter")
            
            soru = akilli_soru_ayikla(soru_metni_blok)
            if soru:
                print(f"✓ Soru {soru_numarasi} başarıyla çıkarıldı")
                print(f"  └─ Soru: {soru['soru_metni'][:60]}...")
                print(f"  └─ Şık sayısı: {sum([1 for k, v in soru.items() if k.startswith('secenek_') and v])}")
                sorular.append(soru)
            else:
                print(f"✗ Soru {soru_numarasi} çıkarılamadı")
            
            i += 2
        
        print(f"\n✅ Toplam çıkarılan soru: {len(sorular)}")
    
    except Exception as e:
        print(f"❌ DOCX işleme hatası: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return sorular


def excel_isle(dosya_yolu):
    """
    Excel dosyasından soru çıkarır - Otomatik kolon algılama
    """
    sorular = []
    
    try:
        workbook = openpyxl.load_workbook(dosya_yolu)
        sheet = workbook.active
        
        # İlk satırı kontrol et - başlık mı?
        # İlk satırı kontrol et - başlık mı?
        baslangic_satir = 1
        ilk_hucre = str(sheet.cell(1, 1).value or '').strip()
        ilk_satir_liste = [str(sheet.cell(1, col).value or '').lower() for col in range(1, 8)]
        
        # Başlık tespiti kriterleri
        baslik_var_mi = False
        kesin_basliklar = ['soru metni', 'soru içeriği', 'soru kökü', 'question text', 'seçenek a', 'option a', 'doğru cevap', 'answer', 'puan']
        
        # Eğer ilk hücre metni 50 karakterden kısaysa başlık olma ihtimali var
        if len(ilk_hucre) < 50:
            # Kesin başlık ifadeleri var mı?
            if any(kb in ' '.join(ilk_satir_liste) for kb in kesin_basliklar):
                baslik_var_mi = True
            # Tam eşleşmeler (Riskli kelimeler için)
            elif ilk_hucre.lower() in ['soru', 'no', 'sıra', 'id']:
                baslik_var_mi = True
                
        if baslik_var_mi:
            baslangic_satir = 2
            print("✓ Excel başlık satırı tespit edildi")
        else:
            print("ℹ Excel ilk satır veri olarak işleniyor (Başlık bulunamadı veya soru metni)")
        
        for satir in range(baslangic_satir, sheet.max_row + 1):
            soru_metni = sheet.cell(satir, 1).value
            
            if not soru_metni or not str(soru_metni).strip():
                continue
            
            soru = {
                'soru_metni': str(soru_metni).strip(),
                'secenek_a': str(sheet.cell(satir, 2).value or '').strip(),
                'secenek_b': str(sheet.cell(satir, 3).value or '').strip(),
                'secenek_c': str(sheet.cell(satir, 4).value or '').strip(),
                'secenek_d': str(sheet.cell(satir, 5).value or '').strip(),
                'secenek_e': str(sheet.cell(satir, 6).value or '').strip(),
                'dogru_cevap': str(sheet.cell(satir, 7).value or 'A').strip().upper(),
                'konu': str(sheet.cell(satir, 8).value or '').strip(),
                'zorluk': str(sheet.cell(satir, 9).value or 'Orta').strip(),
                'puan': int(sheet.cell(satir, 10).value) if sheet.cell(satir, 10).value else 5
            }
            
            sorular.append(soru)
        
        print(f"✅ Excel'den {len(sorular)} soru çıkarıldı")
    
    except Exception as e:
        print(f"❌ Excel işleme hatası: {e}")
        return []
    
    return sorular


def json_isle(dosya_yolu):
    """JSON dosyasından soru çıkarır - Alternatif anahtarları (soru, a, b, cevap vb.) destekler"""
    try:
        with open(dosya_yolu, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        def normalize_soru(s):
            if not isinstance(s, dict):
                return None
                
            # Tuş eşlemeleri (Gelen format -> Veritabanı formatı)
            mapping = {
                'soru': 'soru_metni',
                'a': 'secenek_a',
                'b': 'secenek_b',
                'c': 'secenek_c',
                'd': 'secenek_d',
                'e': 'secenek_e',
                'cevap': 'dogru_cevap'
            }
            
            # Yeni bir kopya oluştur ve eşleşmeleri uygula
            n = s.copy()
            for eski, yeni in mapping.items():
                if eski in s and yeni not in s:
                    n[yeni] = s[eski]
            
            # Zorunlu alanların varlığını ve varsayılanları kontrol et
            if 'soru_metni' not in n or not n['soru_metni']:
                return None
                
            if 'dogru_cevap' not in n or not str(n['dogru_cevap']).strip():
                n['dogru_cevap'] = 'A' # Varsayılan cevap A
            else:
                # Sadece harf kısmını al (Örn: "A)" -> "A")
                match = re.search(r'([A-E])', str(n['dogru_cevap']).upper())
                n['dogru_cevap'] = match.group(1) if match else 'A'

            # Diğer varsayılanlar
            if 'puan' not in n: n['puan'] = 5
            if 'zorluk' not in n: n['zorluk'] = 'Orta'
            if 'gorsel_konum' not in n: n['gorsel_konum'] = 'yanda'
            if 'sik_duzeni' not in n: n['sik_duzeni'] = 'iki_sutun'
            
            return n

        if isinstance(data, dict):
            res = normalize_soru(data)
            return [res] if res else []
        elif isinstance(data, list):
            sorular = []
            for s in data:
                normalized = normalize_soru(s)
                if normalized:
                    sorular.append(normalized)
            return sorular
        else:
            print("⚠ JSON formatı geçersiz: Liste veya Sözlük bekleniyordu.")
            return []
            
    except Exception as e:
        print(f"❌ JSON işleme hatası: {e}")
        return []


def gorsel_isle(dosya_yolu, hedef_klasor):
    """Görsel dosyasını yükler ve yeniden boyutlandırır"""
    try:
        img = Image.open(dosya_yolu)
        
        max_genislik = 800
        if img.width > max_genislik:
            oran = max_genislik / img.width
            yeni_yukseklik = int(img.height * oran)
            img = img.resize((max_genislik, yeni_yukseklik), Image.Resampling.LANCZOS)
        
        dosya_adi = os.path.basename(dosya_yolu)
        hedef_yolu = os.path.join(hedef_klasor, dosya_adi)
        
        img.save(hedef_yolu, optimize=True, quality=85)
        
        return hedef_yolu
    
    except Exception as e:
        print(f"❌ Görsel işleme hatası: {e}")
        return None


def dosya_isle(dosya_yolu, dosya_turu):
    """
    Ana dosya işleme fonksiyonu - Otomatik format algılama
    """
    dosya_turu = dosya_turu.lower()
    
    print(f"🔍 Dosya işleniyor: {os.path.basename(dosya_yolu)} (Tür: {dosya_turu})")
    
    if dosya_turu == 'pdf':
        return pdf_isle(dosya_yolu)
    elif dosya_turu in ['docx', 'doc']:
        return docx_isle(dosya_yolu)
    elif dosya_turu in ['xlsx', 'xls']:
        return excel_isle(dosya_yolu)
    elif dosya_turu == 'json':
        return json_isle(dosya_yolu)
    else:
        print(f"⚠ Desteklenmeyen dosya türü: {dosya_turu}")
        return None
