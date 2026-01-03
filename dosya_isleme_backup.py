"""
Dosya İşleme Modülü - Geliştirilmiş Versiyon
PDF, DOCX ve Excel dosyalarından soru çıkarma işlemleri
"""

import os
import re
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
from PIL import Image


def pdf_isle(dosya_yolu):
    """
    PDF dosyasından soru çıkarır
    
    Beklenen format:
    1. Soru metni burada...
    A) Şık A
    B) Şık B
    C) Şık C
    D) Şık D
    Cevap: A
    """
    sorular = []
    
    try:
        reader = PdfReader(dosya_yolu)
        tum_metin = ""
        
        # Tüm sayfaları oku
        for sayfa in reader.pages:
            tum_metin += sayfa.extract_text() + "\n"
        
        # Soruları ayır (1., 2., 3. gibi numaralandırılmış)
        soru_bloklari = re.split(r'\n\s*\d+\.\s+', tum_metin)
        
        for blok in soru_bloklari[1:]:  # İlk boş bloğu atla
            soru = soru_ayikla_gelismis(blok)
            if soru:
                sorular.append(soru)
    
    except Exception as e:
        print(f"PDF işleme hatası: {e}")
        return []
    
    return sorular


def docx_isle(dosya_yolu):
    """DOCX dosyasından soru çıkarır"""
    sorular = []
    
    try:
        doc = Document(dosya_yolu)
        tum_metin = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        
        print(f"DOCX toplam metin uzunluğu: {len(tum_metin)} karakter")
        
        # Farklı soru formatlarını dene
        # Format 1: "1. Soru..." veya "1) Soru..."
        soru_bloklari = re.split(r'\n\s*(\d+)[.)]\s+', tum_metin)
        
        # Format 2: "S1." veya "Soru 1:" gibi
        if len(soru_bloklari) <= 2:
            soru_bloklari = re.split(r'\n\s*(?:S|Soru)\s*(\d+)[:.)]\s*', tum_metin, flags=re.IGNORECASE)
        
        print(f"Bulunan soru blok sayısı: {len(soru_bloklari)}")
        
        # Blokları işle (1. index numara, 2. index soru metni şeklinde gelir)
        i = 1
        while i < len(soru_bloklari) - 1:
            soru_numarasi = soru_bloklari[i]
            soru_metni_blok = soru_bloklari[i + 1]
            
            print(f"\n--- Soru {soru_numarasi} işleniyor ---")
            print(f"Blok uzunluğu: {len(soru_metni_blok)} karakter")
            
            soru = soru_ayikla_gelismis(soru_metni_blok)
            if soru:
                print(f"✓ Soru {soru_numarasi} başarıyla çıkarıldı")
                print(f"  Soru metni: {soru['soru_metni'][:50]}...")
                print(f"  Şık sayısı: {sum([1 for k, v in soru.items() if k.startswith('secenek_') and v])}")
                sorular.append(soru)
            else:
                print(f"✗ Soru {soru_numarasi} çıkarılamadı")
            
            i += 2  # Numara ve metin çiftlerini atla
        
        print(f"\nToplam çıkarılan soru: {len(sorular)}")
    
    except Exception as e:
        print(f"DOCX işleme hatası: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return sorular


def excel_isle(dosya_yolu):
    """
    Excel dosyasından soru çıkarır
    
    Beklenen format (kolonlar):
    A: Soru Metni
    B: Şık A
    C: Şık B
    D: Şık C
    E: Şık D
    F: Şık E (opsiyonel)
    G: Doğru Cevap (A, B, C, D, E)
    H: Konu (opsiyonel)
    I: Zorluk (opsiyonel)
    J: Puan (opsiyonel)
    """
    sorular = []
    
    try:
        workbook = openpyxl.load_workbook(dosya_yolu)
        sheet = workbook.active
        
        # İlk satır başlık olabilir, kontrol et
        baslangic_satir = 2 if sheet.cell(1, 1).value and 'soru' in str(sheet.cell(1, 1).value).lower() else 1
        
        for satir in range(baslangic_satir, sheet.max_row + 1):
            soru_metni = sheet.cell(satir, 1).value
            
            if not soru_metni or not str(soru_metni).strip():
                continue
            
            soru = {
                'soru_metni': str(soru_metni).strip(),
                'secenek_a': str(sheet.cell(satir, 2).value or '').strip(),
                'secenek_b': str(sheet.cell(satir, 3).value or '').strip(),
                'secenek_c': str(sheet.cell(satir, 4).value or '').strip(),
                'secenek_d': str(sheet.cell(satir, 5).value or '').strip(),
                'secenek_e': str(sheet.cell(satir, 6).value or '').strip(),
                'dogru_cevap': str(sheet.cell(satir, 7).value or 'A').strip().upper(),
                'konu': str(sheet.cell(satir, 8).value or '').strip(),
                'zorluk': str(sheet.cell(satir, 9).value or 'Orta').strip(),
                'puan': int(sheet.cell(satir, 10).value) if sheet.cell(satir, 10).value else 5
            }
            
            sorular.append(soru)
    
    except Exception as e:
        print(f"Excel işleme hatası: {e}")
        return []
    
    return sorular


def gorsel_isle(dosya_yolu, hedef_klasor):
    """
    Görsel dosyasını yükler ve yeniden boyutlandırır
    Döndürür: kaydedilen dosya yolu
    """
    try:
        # Görseli aç
        img = Image.open(dosya_yolu)
        
        # Maksimum genişlik 800px
        max_genislik = 800
        if img.width > max_genislik:
            oran = max_genislik / img.width
            yeni_yukseklik = int(img.height * oran)
            img = img.resize((max_genislik, yeni_yukseklik), Image.Resampling.LANCZOS)
        
        # Dosya adını oluştur
        dosya_adi = os.path.basename(dosya_yolu)
        hedef_yolu = os.path.join(hedef_klasor, dosya_adi)
        
        # Kaydet
        img.save(hedef_yolu, optimize=True, quality=85)
        
        return hedef_yolu
    
    except Exception as e:
        print(f"Görsel işleme hatası: {e}")
        return None


def soru_ayikla_gelismis(metin):
    """
    Gelişmiş soru ayıklama - satır satır şıkları destekler
    """
    try:
        metin = metin.strip()
        if not metin or len(metin) < 10:
            return None
        
        soru = {
            'soru_metni': '',
            'secenek_a': '',
            'secenek_b': '',
            'secenek_c': '',
            'secenek_d': '',
            'secenek_e': '',
            'dogru_cevap': 'A',
            'konu': '',
            'zorluk': 'Orta',
            'puan': 5
        }
        
        satirlar = metin.split('\n')
        soru_metni_satirlari = []
        siklar = {}
        cevap_bulundu = False
        
        for satir in satirlar:
            satir = satir.strip()
            if not satir:
                continue
            
            # Cevap satırını kontrol et
            cevap_match = re.search(r'(?:Cevap|Doğru|Yanıt|Answer)[:\s]*([A-Ea-e])', satir, re.IGNORECASE)
            if cevap_match:
                soru['dogru_cevap'] = cevap_match.group(1).upper()
                cevap_bulundu = True
                continue
            
            # Şık satırını kontrol et - başında A), B), C) vb. olan
            sik_match = re.match(r'^([A-Ea-e])[).:\s]+(.+)$', satir)
            if sik_match:
                harf = sik_match.group(1).upper()
                icerik = sik_match.group(2).strip()
                siklar[harf] = icerik
                continue
            
            # Eğer henüz şık bulunmadıysa, bu soru metninin parçası
            if not siklar and not cevap_bulundu:
                # Puanlama bilgisi içeriyorsa atla
                if re.search(r'\(\d+\s*puan\)', satir, re.IGNORECASE):
                    continue
                soru_metni_satirlari.append(satir)
        
        # Soru metnini birleştir
        soru['soru_metni'] = ' '.join(soru_metni_satirlari).strip()
        
        # Şıkları ata
        soru['secenek_a'] = siklar.get('A', '')
        soru['secenek_b'] = siklar.get('B', '')
        soru['secenek_c'] = siklar.get('C', '')
        soru['secenek_d'] = siklar.get('D', '')
        soru['secenek_e'] = siklar.get('E', '')
        
        # En az soru metni ve bir şık olmalı
        if not soru['soru_metni'] or not soru['secenek_a']:
            print(f"Eksik veri: soru_metni={bool(soru['soru_metni'])}, secenek_a={bool(soru['secenek_a'])}")
            return None
        
        # Doğru cevap bulunamadıysa ve şıklar varsa, içinden tahmin et
        if not cevap_bulundu:
            # Metinde "doğru" kelimesi geçen şıkkı bul
            for harf, icerik in siklar.items():
                if re.search(r'(doğru|correct|✓)', icerik, re.IGNORECASE):
                    soru['dogru_cevap'] = harf
                    break
        
        return soru
    
    except Exception as e:
        print(f"Soru ayıklama hatası: {e}")
        import traceback
        traceback.print_exc()
        return None


def dosya_isle(dosya_yolu, dosya_turu):
    """
    Ana dosya işleme fonksiyonu
    
    Args:
        dosya_yolu: İşlenecek dosyanın yolu
        dosya_turu: Dosya uzantısı (pdf, docx, xlsx, vb.)
    
    Returns:
        Sorular listesi veya None
    """
    dosya_turu = dosya_turu.lower()
    
    if dosya_turu == 'pdf':
        return pdf_isle(dosya_yolu)
    elif dosya_turu in ['docx', 'doc']:
        return docx_isle(dosya_yolu)
    elif dosya_turu in ['xlsx', 'xls']:
        return excel_isle(dosya_yolu)
    else:
        print(f"Desteklenmeyen dosya türü: {dosya_turu}")
        return None
