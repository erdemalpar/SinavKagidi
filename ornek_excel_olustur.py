import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Yeni bir çalışma kitabı oluştur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Soru Bankası"

# Başlık satırı
basliklar = [
    "Soru Metni", 
    "Şık A", 
    "Şık B", 
    "Şık C", 
    "Şık D", 
    "Şık E", 
    "Doğru Cevap", 
    "Konu", 
    "Zorluk", 
    "Puan"
]

# Başlıkları yaz ve formatla
for col, baslik in enumerate(basliklar, start=1):
    hucre = ws.cell(row=1, column=col)
    hucre.value = baslik
    hucre.font = Font(bold=True, color="FFFFFF")
    hucre.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hucre.alignment = Alignment(horizontal="center", vertical="center")

# Örnek sorular ekle
sorular = [
    {
        "soru": "Türkiye'nin başkenti neresidir?",
        "a": "İstanbul",
        "b": "Ankara",
        "c": "İzmir",
        "d": "Bursa",
        "e": "",
        "cevap": "B",
        "konu": "Sosyal Bilgiler",
        "zorluk": "Kolay",
        "puan": 5
    },
    {
        "soru": "2 + 2 işleminin sonucu kaçtır?",
        "a": "2",
        "b": "3",
        "c": "4",
        "d": "5",
        "e": "6",
        "cevap": "C",
        "konu": "Matematik",
        "zorluk": "Kolay",
        "puan": 5
    },
    {
        "soru": "Aşağıdakilerden hangisi bir meyve değildir?",
        "a": "Elma",
        "b": "Armut",
        "c": "Domates",
        "d": "Muz",
        "e": "Portakal",
        "cevap": "C",
        "konu": "Fen Bilgisi",
        "zorluk": "Orta",
        "puan": 10
    },
    {
        "soru": "Osmanlı İmparatorluğu hangi yılda kurulmuştur?",
        "a": "1299",
        "b": "1453",
        "c": "1071",
        "d": "1923",
        "e": "1920",
        "cevap": "A",
        "konu": "Tarih",
        "zorluk": "Orta",
        "puan": 10
    },
    {
        "soru": "Işığın hızı yaklaşık olarak kaç km/s'dir?",
        "a": "150.000 km/s",
        "b": "200.000 km/s",
        "c": "300.000 km/s",
        "d": "400.000 km/s",
        "e": "500.000 km/s",
        "cevap": "C",
        "konu": "Fizik",
        "zorluk": "Zor",
        "puan": 15
    },
    {
        "soru": "Dünya'nın uydusu hangisidir?",
        "a": "Mars",
        "b": "Ay",
        "c": "Güneş",
        "d": "Venüs",
        "e": "Jupiter",
        "cevap": "B",
        "konu": "Fen Bilgisi",
        "zorluk": "Kolay",
        "puan": 5
    },
    {
        "soru": "12 x 5 işleminin sonucu kaçtır?",
        "a": "50",
        "b": "55",
        "c": "60",
        "d": "65",
        "e": "70",
        "cevap": "C",
        "konu": "Matematik",
        "zorluk": "Kolay",
        "puan": 5
    },
    {
        "soru": "Türkiye'nin en uzun nehri hangisidir?",
        "a": "Fırat",
        "b": "Kızılırmak",
        "c": "Dicle",
        "d": "Sakarya",
        "e": "Gediz",
        "cevap": "B",
        "konu": "Coğrafya",
        "zorluk": "Orta",
        "puan": 10
    },
    {
        "soru": "DNA'nın açılımı nedir?",
        "a": "Deoksiribonükleik Asit",
        "b": "Deoksiriboz Nükleik Asam",
        "c": "Dezoksiribonükleik Asit",
        "d": "Dioksiribonükleik Asit",
        "e": "Hiçbiri",
        "cevap": "A",
        "konu": "Biyoloji",
        "zorluk": "Zor",
        "puan": 15
    },
    {
        "soru": "Cumhuriyet hangi yıl ilan edilmiştir?",
        "a": "1920",
        "b": "1921",
        "c": "1922",
        "d": "1923",
        "e": "1924",
        "cevap": "D",
        "konu": "Tarih",
        "zorluk": "Orta",
        "puan": 10
    }
]

# Soruları ekle
for satir, soru in enumerate(sorular, start=2):
    ws.cell(row=satir, column=1).value = soru["soru"]
    ws.cell(row=satir, column=2).value = soru["a"]
    ws.cell(row=satir, column=3).value = soru["b"]
    ws.cell(row=satir, column=4).value = soru["c"]
    ws.cell(row=satir, column=5).value = soru["d"]
    ws.cell(row=satir, column=6).value = soru["e"]
    ws.cell(row=satir, column=7).value = soru["cevap"]
    ws.cell(row=satir, column=8).value = soru["konu"]
    ws.cell(row=satir, column=9).value = soru["zorluk"]
    ws.cell(row=satir, column=10).value = soru["puan"]

# Sütun genişliklerini ayarla
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 10

# Kaydet
dosya_yolu = "/Users/erdem/.gemini/antigravity/scratch/sinav-kagidi/ornek_sorular.xlsx"
wb.save(dosya_yolu)
print(f"Örnek soru dosyası oluşturuldu: {dosya_yolu}")
print(f"Toplam {len(sorular)} soru eklendi.")
